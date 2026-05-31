import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

MOOVA_BLUE = "2563EB"
MOOVA_BLUE_LIGHT = "EFF6FF"
RED_FILL = "FEE2E2"
GREEN_FILL = "DCFCE7"
GRAY_FILL = "F8FAFC"
HEADER_FONT = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=MOOVA_BLUE)
SUBHEADER_FONT = Font(name='Calibri', bold=True, size=9, color="1E3A5F")
SUBHEADER_FILL = PatternFill("solid", fgColor="DBEAFE")
DATA_FONT = Font(name='Calibri', size=9)
BOLD_FONT = Font(name='Calibri', bold=True, size=9)
ERR_FILL = PatternFill("solid", fgColor=RED_FILL)
OK_FILL = PatternFill("solid", fgColor=GREEN_FILL)
thin = Side(style='thin', color='E2E8F0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')


def style_header(ws, row, cols, text=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    if text:
        ws.cell(row=row, column=1).value = text


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def export_ar_control(results, periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Control'

    # Título
    ws.merge_cells('A1:O1')
    ws['A1'] = f'MOOVA · Control de Liquidación Argentina · Período {periodo}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=12, color=MOOVA_BLUE)
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:O2')
    ws['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(name='Calibri', size=9, color='64748B')
    ws['A2'].alignment = CENTER

    # Grupos de columnas
    ws.merge_cells('A4:B4')
    ws.merge_cells('C4:E4')
    ws.merge_cells('F4:I4')
    ws.merge_cells('J4:M4')
    ws.merge_cells('N4:O4')
    for cell, txt in [('A4',''), ('C4','NOVEDADES MOOV'), ('F4','HH.EE CALCULADAS'), ('J4','DATOS ESTUDIO'), ('N4','CONTROL')]:
        ws[cell] = txt
        ws[cell].fill = SUBHEADER_FILL
        ws[cell].font = SUBHEADER_FONT
        ws[cell].alignment = CENTER
        ws[cell].border = BORDER

    # Headers
    headers = ['Leg.','Nombre','HE50%(hs)','HE100%(hs)','Desc.Plan MOOV',
               'Val.Hora','HE50%($)','HE100%($)','CD esp.',
               'Sueldo','CD Estudio','Total Rem.','Neto Estudio',
               'Neto MOOV','DIFERENCIA']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    # Datos
    for row_i, r in enumerate(results, 6):
        hasDiff = abs(r['dif']) > 1
        fill = ERR_FILL if hasDiff else None
        vals = [
            r['leg'], r['nombre'],
            r.get('he50Hs') or '', r.get('he100Hs') or '',
            r.get('descDifMoov') or '',
            r['vh'],
            r.get('he50c') or '', r.get('he100c') or '',
            r.get('cdEsp') or '',
            r['sueldo'], r.get('cdEstudio') or '',
            r['totalRem'], r['neto'],
            r['netoMoov'], r['dif']
        ]
        for col_i, val in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font = DATA_FONT
            c.border = BORDER
            if fill: c.fill = PatternFill("solid", fgColor=RED_FILL)
            c.alignment = RIGHT if col_i > 2 else LEFT

        # Badge estado
        estado = ws.cell(row=row_i, column=15)
        if hasDiff:
            estado.fill = ERR_FILL
            estado.font = Font(name='Calibri', bold=True, size=9, color='DC2626')

    # Totales
    tot_row = len(results) + 6
    ws.cell(row=tot_row, column=2, value='TOTALES').font = BOLD_FONT
    ws.cell(row=tot_row, column=3, value=sum(r.get('he50Hs') or 0 for r in results)).font = BOLD_FONT
    ws.cell(row=tot_row, column=4, value=sum(r.get('he100Hs') or 0 for r in results)).font = BOLD_FONT
    ws.cell(row=tot_row, column=13, value=sum(r['neto'] for r in results)).font = BOLD_FONT
    dif_tot = sum(r['dif'] for r in results)
    c_dif = ws.cell(row=tot_row, column=15, value=dif_tot)
    c_dif.font = Font(name='Calibri', bold=True, size=9,
                      color='DC2626' if abs(dif_tot) > 1 else '16A34A')

    auto_width(ws)
    ws.freeze_panes = 'A6'

    # Hoja diferencias
    diffs = [r for r in results if abs(r['dif']) > 1]
    if diffs:
        ws2 = wb.create_sheet('Diferencias - Reliquidar')
        style_header(ws2, 1, 6, f'Diferencias a reliquidar · {periodo}')
        for i, h in enumerate(['Legajo','Nombre','Neto Estudio','Neto MOOV','Diferencia','Acción'], 1):
            c = ws2.cell(row=2, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        for ri, r in enumerate(diffs, 3):
            ws2.cell(row=ri, column=1, value=r['leg']).font = DATA_FONT
            ws2.cell(row=ri, column=2, value=r['nombre']).font = DATA_FONT
            ws2.cell(row=ri, column=3, value=r['neto']).font = DATA_FONT
            ws2.cell(row=ri, column=4, value=r['netoMoov']).font = DATA_FONT
            c = ws2.cell(row=ri, column=5, value=r['dif'])
            c.font = Font(name='Calibri', bold=True, size=9, color='DC2626')
            ws2.cell(row=ri, column=6, value='RELIQUIDAR').font = Font(name='Calibri', bold=True, size=9, color='DC2626')
        auto_width(ws2)

    # Hoja resumen
    ws3 = wb.create_sheet('Resumen')
    resumen = [
        ('MOOVA · Resumen Control Argentina', ''),
        ('', ''),
        ('Período', periodo),
        ('Fecha de control', datetime.now().strftime('%d/%m/%Y %H:%M')),
        ('', ''),
        ('Total empleados', len(results)),
        ('Sin diferencias', len([r for r in results if not r['hasDiff']])),
        ('Con diferencia', len(diffs)),
        ('', ''),
        ('Neto total estudio', sum(r['neto'] for r in results)),
        ('Diferencia total', sum(r['dif'] for r in results)),
        ('', ''),
        ('HH.EE 50% totales (hs)', sum(r.get('he50Hs') or 0 for r in results)),
        ('HH.EE 100% totales (hs)', sum(r.get('he100Hs') or 0 for r in results)),
    ]
    for ri, (k, v) in enumerate(resumen, 1):
        ws3.cell(row=ri, column=1, value=k).font = BOLD_FONT if k else DATA_FONT
        ws3.cell(row=ri, column=2, value=v).font = DATA_FONT
    auto_width(ws3)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    return tmp.name


def export_mx_novedades(colabs, periodo, manuales=None):
    # Extraer otras novedades si vienen dentro de colabs
    otras_novedades = colabs.pop('__otras_novedades__', manuales or [])
    colabs_clean = {k: v for k, v in colabs.items() if not k.startswith('__')}
    """
    Genera Excel para enviar al estudio MX con 3 hojas:
    1. Novedades - Incidencias: detalle por dia corregido LFT
    2. Resumen: totales por colaborador
    3. Otras novedades: bonos, vacaciones, etc.
    """
    wb = Workbook()

    # ══════════════════════════════════════════════
    # HOJA 1: Novedades - Incidencias
    # ══════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Novedades - Incidencias'

    ws.merge_cells('A1:O1')
    ws['A1'] = f'MOOVA · Novedades para Estudio MX · {periodo}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=11, color=MOOVA_BLUE)
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 22

    headers = [
        '#', 'Colaborador', 'Tipo de Novedad', 'Fecha',
        'Mes de Liquidacion', 'Dia', 'Horario', 'Prima Feriado',
        'Dia Descanso', 'HH EE doble', 'HH EE triple',
        'HE dominical', 'HE Festiva', 'Prima Dominical', 'Comentarios'
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    row_num = 3
    contador = 1
    resumen_data = {}  # para hoja resumen

    for nombre, d in colabs_clean.items():
        filas = d.get('filas', [])
        hasErr = any(a.get('tipo') == 'err' for a in d.get('alertas', []))

        # Acumular para resumen
        resumen_data[nombre] = {
            'heDoble': d.get('heDoble', 0),
            'heTriple': d.get('heTriple', 0),
            'heDom': d.get('heDom', 0),
            'heFest': d.get('heFest', 0),
            'diasDesc': d.get('diasDesc', 0),
            'primaDom': d.get('primaDom', 0),
            'comentarios': d.get('comentarios', ''),
            'totalHE': d.get('totalHE', 0),
            'tieneCorrLFT': bool(d.get('correcciones_lft')),
            'hasErr': hasErr,
        }

        if not filas:
            fill = PatternFill("solid", fgColor=RED_FILL) if hasErr else None
            vals = [contador, nombre, 'Manual', '', '', '', '', '',
                    d.get('diasDesc') or '', d.get('heDoble') or '',
                    d.get('heTriple') or '', d.get('heDom') or '',
                    d.get('heFest') or '', d.get('primaDom') or '',
                    d.get('comentarios') or '']
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font = DATA_FONT; c.border = BORDER
                if fill: c.fill = fill
            row_num += 1; contador += 1
        else:
            for f in filas:
                hasCorrLFT = f.get('tiene_correccion', False)
                if hasErr:
                    fill = PatternFill("solid", fgColor=RED_FILL)
                elif hasCorrLFT:
                    fill = PatternFill("solid", fgColor="FFF9C4")
                else:
                    fill = None

                he_doble = f.get('he_doble', 0)
                he_triple = f.get('he_triple', 0)
                he_dom = f.get('he_dominical', 0)
                he_fest = f.get('he_festiva', 0)
                prima_dom = f.get('prima_dominical', 0)
                dias_desc = f.get('dias_descanso', 0)
                prima_fer = f.get('prima_feriado', 0)

                fecha = str(f.get('fecha', ''))
                if '00:00:00' in fecha:
                    fecha = fecha.replace(' 00:00:00', '')
                    parts = fecha.split('-')
                    if len(parts) == 3:
                        fecha = f"{parts[2]}/{parts[1]}"

                vals = [
                    contador, nombre,
                    f.get('tipo', ''), fecha, f.get('mes', ''),
                    f.get('dia', ''), f.get('horario', ''),
                    prima_fer if prima_fer else '',
                    dias_desc if dias_desc else '',
                    he_doble if he_doble else '',
                    he_triple if he_triple else '',
                    he_dom if he_dom else '',
                    he_fest if he_fest else '',
                    prima_dom if prima_dom else '',
                    f.get('comentarios', '')
                ]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=row_num, column=ci, value=val)
                    c.font = DATA_FONT; c.border = BORDER
                    if fill: c.fill = fill
                row_num += 1; contador += 1

    anchos = [5, 22, 14, 10, 16, 10, 12, 12, 12, 12, 12, 12, 12, 14, 30]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════
    # HOJA 2: Resumen por colaborador
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet('Resumen')

    ws2.merge_cells('A1:I1')
    ws2['A1'] = f'MOOVA · Resumen Novedades · {periodo}'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=11, color=MOOVA_BLUE)
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 22

    hdrs2 = ['Colaborador', 'HE Doble', 'HE Triple', 'HE Dom.', 'HE Festiva',
             'Total HE', 'Días Desc.', 'Prima Dom.', 'Comentarios / Rol']
    for i, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=2, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    tot_heD = tot_heT = tot_heDom = tot_heFest = tot_totalHE = tot_diasDesc = tot_primaDom = 0

    for ri, (nombre, r) in enumerate(resumen_data.items(), 3):
        if r['hasErr']:
            fill = PatternFill("solid", fgColor=RED_FILL)
        elif r['tieneCorrLFT']:
            fill = PatternFill("solid", fgColor="FFF9C4")
        else:
            fill = None

        vals = [
            nombre,
            r['heDoble'] or '',
            r['heTriple'] or '',
            r['heDom'] or '',
            r['heFest'] or '',
            r['totalHE'] or '',
            r['diasDesc'] or '',
            r['primaDom'] or '',
            r['comentarios'] or ''
        ]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = DATA_FONT; c.border = BORDER
            c.alignment = CENTER if ci > 1 else LEFT
            if fill: c.fill = fill

        tot_heD += r['heDoble'] or 0
        tot_heT += r['heTriple'] or 0
        tot_heDom += r['heDom'] or 0
        tot_heFest += r['heFest'] or 0
        tot_totalHE += r['totalHE'] or 0
        tot_diasDesc += r['diasDesc'] or 0
        tot_primaDom += r['primaDom'] or 0

    # Fila de totales
    tot_row = len(resumen_data) + 3
    ws2.cell(row=tot_row, column=1, value='TOTALES').font = BOLD_FONT
    totals = ['', tot_heD or '', tot_heT or '', tot_heDom or '', tot_heFest or '',
              tot_totalHE or '', tot_diasDesc or '', tot_primaDom or '', '']
    for ci, val in enumerate(totals, 1):
        c = ws2.cell(row=tot_row, column=ci, value=val)
        c.font = BOLD_FONT; c.border = BORDER
        c.fill = SUBHEADER_FILL

    ws2.column_dimensions['A'].width = 25
    for col in ['B','C','D','E','F','G','H']:
        ws2.column_dimensions[col].width = 12
    ws2.column_dimensions['I'].width = 30
    ws2.freeze_panes = 'A3'

    # ══════════════════════════════════════════════
    # HOJA 3: Otras novedades
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet('Otras novedades')

    ws3.merge_cells('A1:E1')
    ws3['A1'] = f'MOOVA · Otras Novedades · {periodo}'
    ws3['A1'].font = Font(name='Calibri', bold=True, size=11, color=MOOVA_BLUE)
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 22

    hdrs3 = ['Colaborador', 'Tipo', 'Detalle / Descripción', 'Monto / Cantidad', 'Observaciones']
    for i, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=2, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER

    # Cargar manuales si los hay
    if otras_novedades:
        for ri, m in enumerate(otras_novedades, 3):
            tipo = m.get('tipo', '')
            tipo_label = {
                'bono': 'Bono / Gratificación',
                'ausencia': 'Ausencia / Licencia',
                'cambio_sueldo': 'Cambio de sueldo',
                'vacaciones': 'Vacaciones',
            }.get(tipo, tipo)
            vals = [
                m.get('nombre', ''),
                tipo_label,
                m.get('descripcion', '') or m.get('comentarios', ''),
                m.get('valor', ''),
                m.get('observaciones', '')
            ]
            for ci, val in enumerate(vals, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.font = DATA_FONT; c.border = BORDER
    else:
        # Dejar 10 filas vacías para completar a mano si hace falta
        for ri in range(3, 13):
            for ci in range(1, 6):
                ws3.cell(row=ri, column=ci).border = BORDER

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 35
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 30
    ws3.freeze_panes = 'A3'

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    return tmp.name

def export_mx_control(results, periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Control'

    ws.merge_cells('A1:N1')
    ws['A1'] = f'MOOVA · Control Prenómina México · {periodo}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=12, color=MOOVA_BLUE)
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 24

    headers = ['Clave','Colaborador',
               'HE Doble OPS','HE Doble Est.','Dif.HE Doble',
               'HE Triple OPS','HE Triple Est.','Dif.HE Triple',
               'Días Desc. OPS','Días Desc. Est.','Dif.Días',
               'Percepciones','Neto MXN','Estado']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    for ri, r in enumerate(results, 4):
        hasDiff = r['hasDiff']
        vals = [
            r['clave'], r['nombre'],
            r['heDobleOPS'] if r['heDobleOPS'] is not None else '',
            r['heDobleEst'], r['dHD'] if r['dHD'] is not None else '',
            r['heTripleOPS'] if r['heTripleOPS'] is not None else '',
            r['heTripleEst'], r['dHT'] if r['dHT'] is not None else '',
            r['diasDescOPS'] if r['diasDescOPS'] is not None else '',
            r['diasDescEst'], r['dDD'] if r['dDD'] is not None else '',
            r['totalPerc'], r['neto'],
            'DIFERENCIA' if hasDiff else 'OK'
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DATA_FONT; c.border = BORDER
            if hasDiff: c.fill = PatternFill("solid", fgColor=RED_FILL)

    # Hoja diferencias detalle
    diffs = [r for r in results if r['hasDiff']]
    if diffs:
        ws2 = wb.create_sheet('Diferencias - Corregir')
        for i, h in enumerate(['Clave','Colaborador','Concepto','OPS','Estudio','Diferencia','Acción'], 1):
            c = ws2.cell(row=1, column=i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        ri = 2
        for r in diffs:
            for concepto, ops_v, est_v, dif_v in [
                ('HH.EE Dobles', r['heDobleOPS'], r['heDobleEst'], r['dHD']),
                ('HH.EE Triples', r['heTripleOPS'], r['heTripleEst'], r['dHT']),
                ('Días Descanso', r['diasDescOPS'], r['diasDescEst'], r['dDD']),
            ]:
                if dif_v is not None and abs(dif_v) > 0.1:
                    for ci, val in enumerate([r['clave'], r['nombre'], concepto, ops_v, est_v, dif_v, 'CORREGIR'], 1):
                        c = ws2.cell(row=ri, column=ci, value=val)
                        c.font = DATA_FONT if ci < 7 else Font(name='Calibri', bold=True, size=9, color='DC2626')
                    ri += 1
        auto_width(ws2)

    auto_width(ws)
    ws.freeze_panes = 'A4'

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    return tmp.name
