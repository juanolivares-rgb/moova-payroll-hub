bash

cat > /home/claude/moova_payroll/utils/mx_parser.py << 'PYEOF'
from openpyxl import load_workbook
import re
from datetime import datetime

# ── DÍAS VÁLIDOS PARA PRIMA DOMINICAL ──────────────────────
DIAS_DOMINGO = {'domingo', 'dom', 'do'}
DIAS_SABADO  = {'sabado', 'sábado', 'sab', 'sa'}

# ── REGLAS LFT ─────────────────────────────────────────────
MAX_DOBLES_POR_DIA = 3
MAX_HE_SEMANAL     = 9

def dia_es_domingo(dia_str):
    return str(dia_str or '').lower().strip() in DIAS_DOMINGO

def dia_es_sabado(dia_str):
    return str(dia_str or '').lower().strip() in DIAS_SABADO

def tipo_novedad(t):
    return str(t or '').lower().strip()

def aplicar_reglas_lft(filas):
    """
    Recibe lista de filas del sheet de OPS y aplica las 5 reglas LFT.
    Devuelve lista de filas corregidas + lista de correcciones realizadas.
    """
    corregidas = []
    correcciones = []

    # Agrupar por colaborador para calcular tope semanal
    por_colab = {}
    for f in filas:
        nombre = str(f.get('colaborador', '') or '').strip()
        if nombre not in por_colab:
            por_colab[nombre] = []
        por_colab[nombre].append(f)

    for nombre, rows in por_colab.items():
        he_semana = 0  # acumulador semanal simple

        for f in rows:
            fila = dict(f)
            tipo = tipo_novedad(fila.get('tipo', ''))
            dia  = str(fila.get('dia', '') or '').lower().strip()
            fecha = fila.get('fecha', '')
            comentarios = str(fila.get('comentarios', '') or '').upper()
            he_doble = float(fila.get('he_doble') or 0)
            he_triple = float(fila.get('he_triple') or 0)
            he_dom = float(fila.get('he_dominical') or 0)
            he_fest = float(fila.get('he_festiva') or 0)
            prima_dom = float(fila.get('prima_dominical') or 0)
            dia_desc = float(fila.get('dias_descanso') or 0)

            cambios = []

            # ── REGLA 3: Incapacidad ──────────────────────
            es_incap = 'INCAPACIDAD' in comentarios
            if es_incap and (he_doble > 0 or he_triple > 0 or he_dom > 0):
                cambios.append(f"⚠️ INCAPACIDAD activa — HH.EE ({he_doble}D/{he_triple}T) eliminadas (LFT Art. 42)")
                he_doble = 0
                he_triple = 0
                he_dom = 0
                he_fest = 0

            # ── REGLA 2: Prima dominical en sábado ───────
            if prima_dom > 0 and dia_es_sabado(dia):
                cambios.append(f"⚠️ Prima dominical en SÁBADO eliminada — sábado no es día de descanso dominical (LFT Art. 75)")
                prima_dom = 0

            # ── REGLA 2b: Prima dominical en día hábil ───
            if prima_dom > 0 and not dia_es_domingo(dia) and not dia_es_sabado(dia):
                cambios.append(f"⚠️ Prima dominical en {dia.capitalize()} eliminada — solo aplica en domingo (LFT Art. 75)")
                prima_dom = 0

            # ── REGLA 4: Dominical en sábado → descanso ──
            if ('dominical' in tipo or 'domingo' in tipo) and dia_es_sabado(dia):
                cambios.append(f"⚠️ Tipo '{fila.get('tipo')}' en SÁBADO corregido a Día de Descanso trabajado")
                dia_desc = max(dia_desc, 1)
                prima_dom = 0

            # ── REGLA 1: Tope diario HH.EE dobles (máx 3h/día) ──
            if he_doble > MAX_DOBLES_POR_DIA:
                exceso = he_doble - MAX_DOBLES_POR_DIA
                he_doble_orig = he_doble
                he_doble = MAX_DOBLES_POR_DIA
                he_triple = he_triple + exceso
                cambios.append(f"⚠️ HH.EE dobles {he_doble_orig}h → máx {MAX_DOBLES_POR_DIA}h/día (LFT Art. 68) — {exceso}h convertidas a triples")

            # ── REGLA 1b: Si vienen todas como triples directamente ──
            # (OPS a veces manda todo como triple sin pasar por dobles)
            if he_triple > 0 and he_doble == 0 and tipo == 'horas extra':
                # Las primeras 3 de cada día deberían ser dobles
                # Como no sabemos si son el mismo día, alertamos
                cambios.append(f"⚠️ Verificar: {he_triple}h marcadas como triples directamente — las primeras 3h/día deben ser dobles (LFT Art. 68)")

            # ── REGLA 5: Tope semanal 9h ─────────────────
            total_he_fila = he_doble + he_triple + he_dom + he_fest
            he_semana += total_he_fila
            if he_semana > MAX_HE_SEMANAL:
                cambios.append(f"⚠️ Acumulado semanal {he_semana}h supera tope de {MAX_HE_SEMANAL}h (LFT Art. 68) — verificar distribución")

            fila['he_doble_orig']    = f.get('he_doble') or 0
            fila['he_triple_orig']   = f.get('he_triple') or 0
            fila['prima_dom_orig']   = f.get('prima_dominical') or 0
            fila['he_doble']         = he_doble
            fila['he_triple']        = he_triple
            fila['he_dominical']     = he_dom
            fila['he_festiva']       = he_fest
            fila['prima_dominical']  = prima_dom
            fila['dias_descanso']    = dia_desc
            fila['correcciones_lft'] = cambios
            fila['tiene_correccion'] = len(cambios) > 0

            corregidas.append(fila)
            if cambios:
                correcciones.append({'nombre': nombre, 'fecha': fecha, 'dia': fila.get('dia'), 'cambios': cambios})

    return corregidas, correcciones


def parse_mx_ops(file_obj, manuales=None):
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = list(wb.active.iter_rows(values_only=True))
    tope = 9
    periodo = ''

    r0 = ws[0] if ws else []
    for c in r0:
        s = str(c or '')
        if 'Tope' in s:
            m = re.search(r'(\d+)', s)
            if m: tope = int(m.group(1))
        if '/' in s and len(s) > 8: periodo = s.strip()

    hdr = next((i for i, r in enumerate(ws) if r and r[1] == 'Colaborador'), -1)

    filas_raw = []
    if hdr >= 0:
        for r in ws[hdr + 1:]:
            nombre = str(r[1] or '').strip()
            if not nombre: continue
            filas_raw.append({
                'colaborador': nombre,
                'tipo': str(r[2] or '').strip(),
                'fecha': str(r[3] or ''),
                'mes': str(r[4] or ''),
                'dia': str(r[5] or ''),
                'horario': str(r[6] or ''),
                'prima_feriado': float(r[7] or 0),
                'dias_descanso': float(r[8] or 0),
                'he_doble': float(r[9] or 0),
                'he_triple': float(r[10] or 0),
                'he_dominical': float(r[11] or 0),
                'he_festiva': float(r[12] or 0),
                'prima_dominical': float(r[13] or 0),
                'comentarios': str(r[14] or '') if len(r) > 14 else '',
            })

    # Agregar manuales
    if manuales:
        for m in manuales:
            nombre = str(m.get('nombre', '')).strip()
            if not nombre: continue
            tipo = m.get('tipo', '')
            val = float(m.get('valor', 0))
            fila = {
                'colaborador': nombre,
                'tipo': tipo,
                'fecha': m.get('fecha', ''),
                'mes': '',
                'dia': m.get('dia', ''),
                'horario': '',
                'prima_feriado': 0,
                'dias_descanso': val if tipo == 'diasDesc' else 0,
                'he_doble': val if tipo == 'heDoble' else 0,
                'he_triple': val if tipo == 'heTriple' else 0,
                'he_dominical': val if tipo == 'heDom' else 0,
                'he_festiva': 0,
                'prima_dominical': 0,
                'comentarios': m.get('comentarios', ''),
                'es_manual': True,
                'tipo_manual': tipo,
                'valor_manual': val,
            }
            filas_raw.append(fila)

    # Aplicar reglas LFT
    filas_corregidas, correcciones_lft = aplicar_reglas_lft(filas_raw)

    # Consolidar por colaborador
    colabs = {}
    for f in filas_corregidas:
        nombre = f['colaborador']
        if nombre not in colabs:
            colabs[nombre] = {
                'heDoble': 0, 'heTriple': 0, 'heDom': 0, 'heFest': 0,
                'diasDesc': 0, 'primaDom': 0, 'comentarios': '',
                'alertas': [], 'correcciones_lft': [], 'filas': []
            }
        colabs[nombre]['heDoble']   += f.get('he_doble', 0)
        colabs[nombre]['heTriple']  += f.get('he_triple', 0)
        colabs[nombre]['heDom']     += f.get('he_dominical', 0)
        colabs[nombre]['heFest']    += f.get('he_festiva', 0)
        colabs[nombre]['diasDesc']  += f.get('dias_descanso', 0)
        colabs[nombre]['primaDom']  += f.get('prima_dominical', 0)
        if f.get('comentarios') and not colabs[nombre]['comentarios']:
            colabs[nombre]['comentarios'] = f['comentarios']
        if f.get('correcciones_lft'):
            colabs[nombre]['correcciones_lft'].extend(f['correcciones_lft'])
        colabs[nombre]['filas'].append(f)

    # Alertas por colaborador
    for nombre, d in colabs.items():
        totalHE = d['heDoble'] + d['heTriple'] + d['heDom'] + d['heFest']
        esIncap = 'INCAPACIDAD' in d['comentarios'].upper() or any(
            'INCAPACIDAD' in f.get('comentarios', '').upper() for f in d['filas'])
        if esIncap and totalHE > 0:
            d['alertas'].append({'tipo': 'err', 'msg': 'HH.EE registradas durante incapacidad — no liquidable (LFT Art. 42)'})
        if totalHE > tope * 2:
            d['alertas'].append({'tipo': 'warn', 'msg': f'Total HH.EE ({totalHE}h) puede superar tope semanal de {tope}h'})
        d['totalHE'] = totalHE
        # Agregar correcciones LFT como alertas info
        if d['correcciones_lft']:
            d['alertas'].append({'tipo': 'info', 'msg': f'{len(d["correcciones_lft"])} corrección(es) LFT aplicada(s)'})

    return colabs, periodo, tope, correcciones_lft


def parse_mx_prenomina(file_obj):
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws_name = next((s for s in wb.sheetnames if 'NOMINA' in s.upper()), wb.sheetnames[0])
    rows = list(wb[ws_name].iter_rows(values_only=True))
    empleados = {}
    emp_actual = None
    periodo = ''

    for r in rows:
        c0 = str(r[0] or '').strip()
        if 'Período del' in c0: periodo = c0.strip()
        if c0 and c0.isdigit() and len(c0) == 6:
            nombre = ' '.join(str(r[1] or '').split())
            emp_actual = nombre
            empleados[nombre] = {
                'clave': c0,
                'totalPerc': float(r[5] or 0),
                'totalISR': float(r[7] or 0),
                'totalDesc': float(r[8] or 0),
                'neto': float(r[9] or 0),
                'conceptos': []
            }
        elif emp_actual and r[5] and str(r[5]).strip().isdigit():
            empleados[emp_actual]['conceptos'].append({
                'cod': str(r[5]).strip(),
                'desc': str(r[6] or '').strip(),
                'cantidad': r[7],
                'perc': float(r[8] or 0) if r[8] else None,
                'ded': float(r[9] or 0) if r[9] else None
            })

    return empleados, periodo


def run_mx_control(colabs, empleados):
    def norm(s): return ' '.join(s.upper().split())

    results = []
    for nombreEst, emp in empleados.items():
        opsKey = next((k for k in colabs if any(
            w in norm(nombreEst) for w in norm(k).split() if len(w) > 3)), None)
        ops = colabs.get(opsKey)

        def getCant(cod):
            return sum(float(c['cantidad'] or 0) for c in emp['conceptos']
                       if c['cod'] == cod and c['cantidad'])

        heDobleEst  = getCant('028')
        heTripleEst = getCant('029')
        diasDescEst = getCant('030')
        primaDomEst = getCant('033')

        heDobleOPS  = ops['heDoble']  if ops else None
        heTripleOPS = ops['heTriple'] if ops else None
        diasDescOPS = ops['diasDesc'] if ops else None
        primaDomOPS = ops['primaDom'] if ops else None

        dHD = round(heDobleEst  - (heDobleOPS  or 0), 2) if heDobleOPS  is not None else None
        dHT = round(heTripleEst - (heTripleOPS or 0), 2) if heTripleOPS is not None else None
        dDD = round(diasDescEst - (diasDescOPS or 0), 2) if diasDescOPS is not None else None
        dPD = round(primaDomEst - (primaDomOPS or 0), 2) if primaDomOPS is not None else None

        hasDiff = any(d is not None and abs(d) > 0.1 for d in [dHD, dHT, dDD, dPD])
        esIncap  = ops and any('incapacidad' in a['msg'].lower() for a in ops.get('alertas', []))
        alertaIncap = esIncap and (heDobleEst > 0 or heTripleEst > 0)

        results.append({
            'nombre': nombreEst, 'clave': emp['clave'], 'opsMatch': opsKey,
            'heDobleOPS': heDobleOPS, 'heTripleOPS': heTripleOPS,
            'diasDescOPS': diasDescOPS, 'primaDomOPS': primaDomOPS,
            'heDobleEst': heDobleEst, 'heTripleEst': heTripleEst,
            'diasDescEst': diasDescEst, 'primaDomEst': primaDomEst,
            'dHD': dHD, 'dHT': dHT, 'dDD': dDD, 'dPD': dPD,
            'totalPerc': emp['totalPerc'], 'totalDesc': emp['totalDesc'],
            'neto': emp['neto'],
            'hasDiff': hasDiff, 'alertaIncap': alertaIncap,
        })

    return sorted(results, key=lambda x: (-x['hasDiff'], x['nombre']))
PYEOF
echo "mx_parser.py OK - $(wc -l < /home/claude/moova_payroll/utils/mx_parser.py) líneas"
Salida

mx_parser.py OK - 307 líneas
