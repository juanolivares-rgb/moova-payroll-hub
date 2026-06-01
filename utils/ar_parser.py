from openpyxl import load_workbook

DIAS_100 = {'sa', 'sab', 'sábado', 'sabado', 'do', 'dom', 'domingo', 'feriado', 'fer'}

def clasificar_he(dia_semana):
    d = str(dia_semana or '').lower().strip()
    return 100 if any(x in d for x in DIAS_100) else 50

def parse_ar_novedades(file_obj, manuales=None):
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    novMap = {}

    # Hoja Carga novedades — HH.EE
    if 'Carga novedades' in wb.sheetnames:
        ws = list(wb['Carga novedades'].iter_rows(values_only=True))
        hdr = next((i for i, r in enumerate(ws) if r and r[0] == 'Legajo'), -1)
        if hdr >= 0:
            for r in ws[hdr + 1:]:
                if not r[0]: continue
                try: leg = int(float(r[0]))
                except: continue
                cod = str(r[2] or '').strip()
                val = float(r[4] or 0)
                if leg not in novMap:
                    novMap[leg] = {'he50': 0, 'he100': 0, 'descDifPlan': 0}
                if 'RE001020' in cod: novMap[leg]['he50'] += val
                if 'RE001021' in cod: novMap[leg]['he100'] += val

    # Hoja Carga Manual — descuentos
    if 'Carga Manual' in wb.sheetnames:
        ws = list(wb['Carga Manual'].iter_rows(values_only=True))
        hdr = next((i for i, r in enumerate(ws) if r and r[0] == 'Legajo'), -1)
        if hdr >= 0:
            for r in ws[hdr + 1:]:
                if not r[0]: continue
                try: leg = int(float(r[0]))
                except: continue
                val = float(r[3] or 0)
                if leg not in novMap:
                    novMap[leg] = {'he50': 0, 'he100': 0, 'descDifPlan': 0}
                novMap[leg]['descDifPlan'] += val

    # Período
    ws_first = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    periodo = ''
    for row in ws_first:
        for c in row:
            s = str(c or '')
            if s.count('/') == 1 and len(s) == 7:
                periodo = s; break
        if periodo: break

    # Novedades manuales (formulario)
    if manuales:
        for m in manuales:
            try: leg = int(m.get('legajo', 0))
            except: continue
            if not leg: continue
            if leg not in novMap:
                novMap[leg] = {'he50': 0, 'he100': 0, 'descDifPlan': 0}
            tipo = m.get('tipo', '')
            val = float(m.get('valor', 0))
            if tipo == 'he50': novMap[leg]['he50'] += val
            elif tipo == 'he100': novMap[leg]['he100'] += val
            elif tipo == 'descuento': novMap[leg]['descDifPlan'] += val
            elif tipo == 'bono': novMap[leg].setdefault('bono', 0); novMap[leg]['bono'] += val
            elif tipo == 'ausencia': novMap[leg].setdefault('ausencia', 0); novMap[leg]['ausencia'] += val

    return novMap, periodo


def parse_ar_ops_imagen(registros):
    """Convierte registros extraídos por Claude Vision al formato novMap."""
    novMap = {}
    for r in registros:
        nombre = r.get('nombre', '')
        dia = r.get('dia_semana', '')
        horas = float(r.get('horas', 0))
        tipo = clasificar_he(dia)
        key = nombre
        if key not in novMap:
            novMap[key] = {'nombre': nombre, 'he50': 0, 'he100': 0, 'filas': []}
        if tipo == 50:
            novMap[key]['he50'] += horas
        else:
            novMap[key]['he100'] += horas
        novMap[key]['filas'].append(r)
    return novMap


def parse_ar_liquidacion(file_obj):
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws_name = next((s for s in wb.sheetnames if 'MENSUAL' in s.upper()), wb.sheetnames[0])
    ws = list(wb[ws_name].iter_rows(values_only=True))
    if not ws: return {}

    headers = [str(h or '').strip() for h in ws[0]]
    def col(p): return next((i for i, h in enumerate(headers) if h.startswith(p)), None)

    CI = {
        'leg': col('LEGAJO') or 1, 'nom': col('APELLIDO') or 2,
        'sdo': col('RE000001'), 'cd': col('RE000561'),
        'tr': col('TOTALREM'), 'dl': col('TOTALDESCLEY'),
        'da': col('DE000010'), 'ddp': col('DE000020'),
        'tod': col('TOTALOTRDESC'), 'net': col('NETO')
    }

    empMap = {}
    for r in ws[1:]:
        try:
            leg_raw = r[CI['leg']] if CI['leg'] is not None and CI['leg'] < len(r) else None
        except (IndexError, TypeError):
            continue
        if not leg_raw: continue
        try: leg = int(float(leg_raw))
        except: continue
        def safe_get(row, col_idx):
            if col_idx is None or col_idx >= len(row): return 0
            try: return float(row[col_idx] or 0)
            except: return 0

        empMap[leg] = {
            'nombre': str(r[CI['nom']] or '') if CI['nom'] is not None and CI['nom'] < len(r) else '',
            'sueldo': safe_get(r, CI['sdo']),
            'compDisp': safe_get(r, CI['cd']),
            'totalRem': safe_get(r, CI['tr']),
            'descLey': safe_get(r, CI['dl']),
            'descAntic': safe_get(r, CI['da']),
            'descDifPlan': safe_get(r, CI['ddp']),
            'totalOtrDesc': safe_get(r, CI['tod']),
            'neto': safe_get(r, CI['net']),
        }
    return empMap


def run_ar_control(novMap, empMap):
    results = []
    for leg, emp in empMap.items():
        nov = novMap.get(leg, {'he50': 0, 'he100': 0, 'descDifPlan': 0})
        vh = emp['sueldo'] / 200 if emp['sueldo'] else 0
        he50c = round(vh * 1.5 * (nov.get('he50') or 0), 2)
        he100c = round(vh * 2 * (nov.get('he100') or 0), 2)
        cdEsp = he50c + he100c
        tieneHE = (nov.get('he50') or 0) > 0 or (nov.get('he100') or 0) > 0
        descFinal = (nov.get('descDifPlan') or 0) if (nov.get('descDifPlan') or 0) > 0 else emp['descDifPlan']
        desvioCD = (cdEsp - emp['compDisp']) if tieneHE else 0
        netoMoov = emp['totalRem'] - emp['descLey'] - emp['descAntic'] - descFinal + desvioCD
        dif = round(netoMoov - emp['neto'], 2)
        difCD = round(desvioCD, 2) if tieneHE else None
        results.append({
            'leg': leg,
            'nombre': emp['nombre'],
            'he50Hs': nov.get('he50') or None,
            'he100Hs': nov.get('he100') or None,
            'descDifMoov': nov.get('descDifPlan') or None,
            'vh': round(vh, 4),
            'he50c': he50c or None,
            'he100c': he100c or None,
            'cdEsp': cdEsp or None,
            'sueldo': emp['sueldo'],
            'cdEstudio': emp['compDisp'] or None,
            'totalRem': emp['totalRem'],
            'descLey': emp['descLey'],
            'neto': emp['neto'],
            'netoMoov': round(netoMoov, 2),
            'difCD': difCD,
            'dif': dif,
            'hasDiff': abs(dif) > 1,
        })
    return sorted(results, key=lambda x: (-abs(x['dif']), x['leg']))
