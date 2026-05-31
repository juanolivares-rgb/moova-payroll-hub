import os
import json
import sqlite3
import base64
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, g
from openpyxl import load_workbook
import anthropic
from dotenv import load_dotenv
from utils.ar_parser import parse_ar_novedades, parse_ar_liquidacion, run_ar_control
from utils.mx_parser import parse_mx_ops, parse_mx_prenomina, run_mx_control
from utils.exporters import export_ar_control, export_mx_novedades, export_mx_control
from utils.db import init_db, save_period_ar, save_period_mx, get_metrics

load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

ANTHROPIC_KEY = os.getenv('ANTHROPIC_API_KEY')

# ── DB ──────────────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect('moova_payroll.db')
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

with app.app_context():
    init_db(sqlite3.connect('moova_payroll.db'))

# ── ROUTES ──────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/ar')
def ar():
    return render_template('ar.html')

@app.route('/mx')
def mx():
    return render_template('mx.html')

@app.route('/metrics')
def metrics():
    return render_template('metrics.html')

# ── API: IMAGEN → CLAUDE VISION ─────────────────────────────
@app.route('/api/ar/export-novedades', methods=['POST'])
def ar_export_novedades():
    data = request.get_json()
    novedades = data.get('novedades', {})
    periodo = data.get('periodo', 'PERIODO')
    
    # Convertir formato imagen a formato exportable
    rows = []
    for nombre, d in novedades.items():
        rows.append({
            'nombre': nombre,
            'he50': d.get('he50', 0),
            'he100': d.get('he100', 0),
            'totalHE': d.get('he50', 0) + d.get('he100', 0)
        })
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Novedades HR Strategy'
    
    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = f'MOOVA · Novedades para HR Strategy · {periodo}'
    ws['A1'].font = Font(bold=True, size=12, color='2563EB')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Apellido y Nombre', 'HH.EE 50% (hs)', 'HH.EE 100% (hs)', 'Total HH.EE']
    blue_fill = PatternFill("solid", fgColor="2563EB")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = blue_fill
        c.alignment = Alignment(horizontal='center')
    
    for ri, r in enumerate(rows, 4):
        ws.cell(row=ri, column=1, value=r['nombre'])
        ws.cell(row=ri, column=2, value=r['he50'] or '')
        ws.cell(row=ri, column=3, value=r['he100'] or '')
        ws.cell(row=ri, column=4, value=r['totalHE'])
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    
    per = periodo.replace('/', '-') if periodo else 'PERIODO'
    return send_file(tmp.name, as_attachment=True,
                     download_name=f'AR-novedades-HR-Strategy-{per}.xlsx')
def extract_image():
    if 'image' not in request.files:
        return jsonify({'error': 'No image uploaded'}), 400
    
    img = request.files['image']
    country = request.form.get('country', 'ar')
    img_data = base64.standard_b64encode(img.read()).decode('utf-8')
    media_type = img.content_type or 'image/png'
    
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    
    if country == 'ar':
        prompt = """Analizá esta planilla de horas extra de Argentina. 
        Extraé TODOS los registros y devolvé SOLO un JSON con este formato exacto:
        {
          "registros": [
            {
              "nombre": "Nombre Apellido",
              "puesto": "Puesto",
              "fecha": "DD/MM",
              "mes": "Mes",
              "dia_semana": "Lu/Ma/Mi/Ju/Vi/Sa/Do",
              "horario": "HH a HH",
              "horas": número
            }
          ]
        }
        El tipo de HH.EE se clasifica automáticamente: Sa/Do/Feriado = 100%, resto = 50%.
        Devolvé SOLO el JSON, sin texto adicional."""
    else:
        prompt = """Analizá esta planilla de novedades de México (HH.EE).
        Extraé TODOS los registros y devolvé SOLO un JSON con este formato exacto:
        {
          "registros": [
            {
              "colaborador": "Nombre Apellido",
              "tipo": "tipo de jornada",
              "fecha": "DD/MM",
              "dia_semana": "nombre del día",
              "horario": "HH a HH",
              "he_doble": número,
              "he_triple": número,
              "he_dominical": número,
              "he_festiva": número,
              "dias_descanso": número,
              "prima_dominical": número,
              "comentarios": "texto o vacío"
            }
          ]
        }
        Devolvé SOLO el JSON, sin texto adicional."""
    
    try:
        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=4000,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_data}},
                    {"type": "text", "text": prompt}
                ]
            }]
        )
        raw = response.content[0].text.strip()
        # Limpiar posibles ```json
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1].rsplit('```', 1)[0]
        data = json.loads(raw)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: AR CONTROL ─────────────────────────────────────────
@app.route('/api/ar/control', methods=['POST'])
def ar_control():
    nov_file = request.files.get('novedades')
    liq_file = request.files.get('liquidacion')
    manuales = request.form.get('manuales', '[]')
    
    if not nov_file or not liq_file:
        return jsonify({'error': 'Faltan archivos'}), 400
    
    try:
        manuales_data = json.loads(manuales)
        nov_data, periodo = parse_ar_novedades(nov_file, manuales_data)
        liq_data = parse_ar_liquidacion(liq_file)
        results = run_ar_control(nov_data, liq_data)
        
        # Guardar en métricas
        db = get_db()
        save_period_ar(db, periodo, results)
        
        return jsonify({'success': True, 'periodo': periodo, 'results': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: AR EXPORT ──────────────────────────────────────────
@app.route('/api/ar/export', methods=['POST'])
def ar_export():
    data = request.get_json()
    periodo = data.get('periodo', 'PERIODO')
    results = data.get('results', [])
    filepath = export_ar_control(results, periodo)
    return send_file(filepath, as_attachment=True,
                     download_name=f'AR-control-nomina-{periodo.replace("/","-")}.xlsx')

# ── API: AR MENSAJE ESTUDIO (CLAUDE) ────────────────────────
@app.route('/api/ar/mensaje-estudio', methods=['POST'])
def ar_mensaje_estudio():
    data = request.get_json()
    diffs = data.get('diffs', [])
    periodo = data.get('periodo', '')
    
    if not diffs:
        return jsonify({'mensaje': ''})
    
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    
    diff_txt = '\n'.join([
        f"- {d['nombre']} (Leg.{d['leg']}): Neto estudio ${d['neto']:,.2f} vs MOOV ${d['netoMoov']:,.2f} → Dif: ${d['dif']:+,.2f}"
        for d in diffs
    ])
    
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=800,
        messages=[{"role": "user", "content": f"""
Redactá un mensaje profesional y cordial para enviar a Edith (HR Strategy), el estudio de liquidación de Argentina.
El mensaje debe informar sobre diferencias encontradas en la preliquidación del período {periodo} y solicitar corrección.

Diferencias detectadas:
{diff_txt}

Tono: formal pero directo, sin introducciones largas. Firmá como Juan.
Devolvé SOLO el mensaje, sin explicaciones adicionales.
"""}]
    )
    return jsonify({'mensaje': response.content[0].text})

# ── API: AR RESPUESTA EMPLEADO (CLAUDE) ─────────────────────
@app.route('/api/ar/respuesta-empleado', methods=['POST'])
def ar_respuesta_empleado():
    data = request.get_json()
    empleado = data.get('empleado', {})
    pregunta = data.get('pregunta', '')
    
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    
    concepto_txt = f"""
Nombre: {empleado.get('nombre')}
Sueldo básico: ${empleado.get('sueldo', 0):,.2f}
HH.EE 50%: {empleado.get('he50Hs') or 0} hs (${empleado.get('he50c') or 0:,.2f})
HH.EE 100%: {empleado.get('he100Hs') or 0} hs (${empleado.get('he100c') or 0:,.2f})
Total remuneración: ${empleado.get('totalRem', 0):,.2f}
Descuentos de ley: ${empleado.get('descLey', 0):,.2f}
Neto: ${empleado.get('neto', 0):,.2f}
"""
    
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content": f"""
Sos el área de People de MOOVA. Respondé la consulta de este colaborador sobre su recibo de sueldo.

Datos del recibo:
{concepto_txt}

Consulta del colaborador: "{pregunta}"

Tono: claro, humano, sin tecnicismos. Explicá los números de forma simple.
Devolvé SOLO la respuesta para el empleado.
"""}]
    )
    return jsonify({'respuesta': response.content[0].text})

# ── API: MX NOVEDADES ───────────────────────────────────────
@app.route('/api/mx/novedades', methods=['POST'])
def mx_novedades():
    ops_file = request.files.get('ops')
    manuales = request.form.get('manuales', '[]')
    
    if not ops_file:
        return jsonify({'error': 'Falta archivo de OPS'}), 400
    
    try:
        manuales_data = json.loads(manuales)
        colabs, periodo, tope = parse_mx_ops(ops_file, manuales_data)
        return jsonify({'success': True, 'periodo': periodo, 'tope': tope, 'colabs': colabs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: MX EXPORT NOVEDADES ────────────────────────────────
@app.route('/api/mx/export-novedades', methods=['POST'])
def mx_export_novedades():
    data = request.get_json()
    colabs = data.get('colabs', {})
    periodo = data.get('periodo', 'PERIODO')
    filepath = export_mx_novedades(colabs, periodo)
    return send_file(filepath, as_attachment=True,
                     download_name=f'MX-novedades-estudio-{periodo.replace("/","-").replace(" ","-")}.xlsx')

# ── API: MX CONTROL ─────────────────────────────────────────
@app.route('/api/mx/control', methods=['POST'])
def mx_control():
    ops_file = request.files.get('ops')
    nom_file = request.files.get('nomina')
    manuales = request.form.get('manuales', '[]')
    
    if not nom_file:
        return jsonify({'error': 'Falta prenómina del estudio'}), 400
    
    try:
        manuales_data = json.loads(manuales)
        colabs, periodo, tope = parse_mx_ops(ops_file, manuales_data) if ops_file else ({}, '', 9)
        empleados, periodo_nom = parse_mx_prenomina(nom_file)
        results = run_mx_control(colabs, empleados)
        per = periodo_nom or periodo
        
        db = get_db()
        save_period_mx(db, per, results, colabs)
        
        return jsonify({'success': True, 'periodo': per, 'results': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── API: MX MENSAJE ESTUDIO (CLAUDE) ────────────────────────
@app.route('/api/mx/mensaje-estudio', methods=['POST'])
def mx_mensaje_estudio():
    data = request.get_json()
    diffs = data.get('diffs', [])
    periodo = data.get('periodo', '')
    
    if not diffs:
        return jsonify({'mensaje': ''})
    
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    
    diff_txt = '\n'.join([
        f"- {d['nombre']}: HE dobles OPS={d.get('heDobleOPS','?')} vs Estudio={d.get('heDobleEst',0)}, "
        f"HE triples OPS={d.get('heTripleOPS','?')} vs Estudio={d.get('heTripleEst',0)}, "
        f"Días descanso OPS={d.get('diasDescOPS','?')} vs Estudio={d.get('diasDescEst',0)}"
        for d in diffs
    ])
    
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=800,
        messages=[{"role": "user", "content": f"""
Redactá un mensaje formal para el estudio de nómina de México (Estudio MX / RC Abogados).
Informá sobre diferencias entre las novedades enviadas y la prenómina del período {periodo}.
Solicitá corrección y reenvío.

Diferencias:
{diff_txt}

Tono: formal, cordial, directo. Firmá como Juan.
Devolvé SOLO el mensaje.
"""}]
    )
    return jsonify({'mensaje': response.content[0].text})

# ── API: MX EXPORT CONTROL ──────────────────────────────────
@app.route('/api/mx/export-control', methods=['POST'])
def mx_export_control():
    data = request.get_json()
    results = data.get('results', [])
    periodo = data.get('periodo', 'PERIODO')
    filepath = export_mx_control(results, periodo)
    return send_file(filepath, as_attachment=True,
                     download_name=f'MX-control-nomina-{periodo.replace("/","-").replace(" ","-")}.xlsx')

# ── API: MÉTRICAS ────────────────────────────────────────────
@app.route('/api/metrics', methods=['GET'])
def api_metrics():
    try:
        db = get_db()
        data = get_metrics(db)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
